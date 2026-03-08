import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Border, Side

# A cor escolhida em formato Hexadecimal (RGB: 31, 73, 125)
COR_PRIMARIA_HEX = "1F497D"

def aplicar_estilo_ancora(worksheet, df):
    """
    Varre a aba do Excel recém-criada e aplica a identidade visual corporativa:
    - Cabeçalho: Fundo Azul Escuro, Fonte Branca em Negrito.
    - Linhas de Dados: Bordas finas na mesma cor do cabeçalho.
    """
    estilo_fundo_cabecalho = PatternFill(start_color=COR_PRIMARIA_HEX, end_color=COR_PRIMARIA_HEX, fill_type="solid")
    estilo_fonte_cabecalho = Font(bold=True, color="FFFFFF") 
    
    linha_borda = Side(border_style="thin", color=COR_PRIMARIA_HEX)
    estilo_borda_celula = Border(left=linha_borda, right=linha_borda, top=linha_borda, bottom=linha_borda)
    
    max_row = df.shape[0] + 1 
    max_col = df.shape[1]
    
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = worksheet.cell(row=r, column=c)
            cell.border = estilo_borda_celula
            
            if r == 1:
                cell.fill = estilo_fundo_cabecalho
                cell.font = estilo_fonte_cabecalho

def aplicar_estilo_basico(worksheet, df):
    """
    Aplica um formato cru/neutro para extrações de banco de dados:
    - Retira o negrito do cabeçalho.
    - Força a remoção de TODAS as bordas (sobrepondo o padrão do Pandas).
    """
    estilo_fonte_normal = Font(bold=False) # Força a remoção do negrito do Pandas
    sem_borda = Border() # Borda vazia (None) atua como uma borracha
    
    max_row = df.shape[0] + 1
    max_col = df.shape[1]
    
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = worksheet.cell(row=r, column=c)
            
            # Aplica a borracha em TODAS as células para matar qualquer padrão
            cell.border = sem_borda
            
            # Tira o negrito apenas da primeira linha
            if r == 1:
                cell.font = estilo_fonte_normal

# ==========================================
# FUNÇÕES PÚBLICAS DE EXPORTAÇÃO
# ==========================================

def exportar_devolutiva_erros(df_dados):
    """Gera o arquivo de erros/críticas (Passo 4 da Análise) com o estilo da empresa."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        nome_aba = 'Devolutiva_Erros'
        df_dados.to_excel(writer, sheet_name=nome_aba, index=False)
        worksheet = writer.sheets[nome_aba]
        aplicar_estilo_ancora(worksheet, df_dados)
        
    return output.getvalue()

def exportar_lista_limpa(df_principal, df_excluidos):
    """Gera o arquivo final da lista de preços (Passo 3 ou 5) separando o lixo em outra aba."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # Aba 1: Lista Consolidada
        df_principal.to_excel(writer, sheet_name='Lista_Consolidada', index=False)
        worksheet1 = writer.sheets['Lista_Consolidada']
        aplicar_estilo_ancora(worksheet1, df_principal)
        
        # Aba 2: Lixo / Excluídos
        if not df_excluidos.empty:
            df_excluidos.to_excel(writer, sheet_name='Excluidos_Automaticamente', index=False)
            worksheet2 = writer.sheets['Excluidos_Automaticamente']
            aplicar_estilo_ancora(worksheet2, df_excluidos)
        else:
            df_vazio = pd.DataFrame([{"Mensagem": "Nenhuma linha precisou ser excluída automaticamente."}])
            df_vazio.to_excel(writer, sheet_name='Excluidos_Automaticamente', index=False)
            worksheet2 = writer.sheets['Excluidos_Automaticamente']
            aplicar_estilo_ancora(worksheet2, df_vazio)
            
    return output.getvalue()

def exportar_consulta_sql(df_dados):
    """Gera a extração bruta do banco de dados em um layout neutro."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        nome_aba = 'Resultado_Consulta'
        df_dados.to_excel(writer, sheet_name=nome_aba, index=False)
        
        # Aplica o estilo "clean" (sem cores, sem negrito)
        worksheet = writer.sheets[nome_aba]
        aplicar_estilo_basico(worksheet, df_dados)
        
    return output.getvalue()